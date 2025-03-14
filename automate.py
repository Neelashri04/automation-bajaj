import pandas as pd
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
import os
import warnings
import streamlit as st
import io
import tempfile

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def process_excel(input_file, output_file, reference_file=None):
    # Read the input Excel file
    df = pd.read_excel(input_file)

    # Use the specific column name "MutualFund Name" for the input file
    fund_name_col = 'MutualFund Name'
    if fund_name_col not in df.columns:
        st.error(f"Error: Could not find column '{fund_name_col}' in the input file.")
        return False

    # Identify rows containing 'total' (case-insensitive)
    total_rows = df[df[fund_name_col].str.contains('total', case=False, na=False)]
    df = df[~df[fund_name_col].str.contains('total', case=False, na=False)]

    # Load reference file if it exists
    reference_data = {}
    reference_sheet_d1_values = {}
    if reference_file:
        try:
            # Load all sheets from the reference file
            reference_wb = load_workbook(reference_file, read_only=True)
            
            # First, extract D1 cell values from each sheet
            for sheet_name in reference_wb.sheetnames:
                sheet = reference_wb[sheet_name]
                d1_value = sheet['D1'].value if sheet['D1'].value else ""
                reference_sheet_d1_values[sheet_name] = str(d1_value).strip()
            
            # Now process each sheet for mutual fund and rank data
            for sheet_name in reference_wb.sheetnames:
                # Create a DataFrame for each sheet
                sheet_df = pd.read_excel(reference_file, sheet_name=sheet_name)
                
                # Find columns containing "mutual fund" (case-insensitive) and exact match for "Rank"
                ref_fund_name_col = None
                ref_rank_col = None
                for col in sheet_df.columns:
                    col_str = str(col)
                    if 'mutual fund' in col_str.lower():
                        ref_fund_name_col = col
                    elif col_str == 'Rank':  # Exact match for "Rank"
                        ref_rank_col = col
                
                # Check if the DataFrame has the expected columns
                if ref_rank_col and ref_fund_name_col:
                    # Store the rank and mutual fund name mapping
                    reference_data[sheet_name] = {
                        str(name): rank for name, rank in zip(sheet_df[ref_fund_name_col], sheet_df[ref_rank_col])
                        if pd.notna(name) and pd.notna(rank)
                    }
        except Exception as e:
            st.error(f"Error reading reference file: {e}")
            return False

    # Create a new Excel workbook
    wb = Workbook()

    # Iterate over each column (except the first one and the fund name column)
    value_columns = [col for col in df.columns if col != df.columns[0] and col != fund_name_col]
    for column in value_columns:
        # Filter rows where the column has a valid number
        filtered_df = df[[fund_name_col, column]].dropna()
        filtered_df = filtered_df[filtered_df[column].apply(lambda x: isinstance(x, (int, float)))]
        # Sort the data in descending order
        filtered_df = filtered_df.sort_values(by=column, ascending=False)
        # Add a 'Rank' column
        filtered_df['Rank'] = range(1, len(filtered_df) + 1)
        # Move 'Rank' to the first column
        cols = ['Rank'] + [col for col in filtered_df.columns if col != 'Rank']
        filtered_df = filtered_df[cols]
        
        # Rename the fund name column to "Mutual Fund"
        filtered_df = filtered_df.rename(columns={fund_name_col: 'Mutual Fund'})

        # Add a 'Rank Change' column (as numeric)
        filtered_df['Rank Change'] = 0
        
        # Check for rank changes if reference data exists
        sheet_name = column[:31]
        matching_sheet = None
        
        if reference_data:
            # Find a matching sheet in the reference data by checking D1 cell value
            for ref_sheet, d1_value in reference_sheet_d1_values.items():
                # Check if the D1 cell value matches the current sheet name (case-insensitive)
                if d1_value.lower() == sheet_name.lower() or sheet_name.lower() in d1_value.lower() or d1_value.lower() in sheet_name.lower():
                    matching_sheet = ref_sheet
                    break
            
            if matching_sheet and matching_sheet in reference_data:
                # Compare ranks for each mutual fund
                for idx, row in filtered_df.iterrows():
                    fund_name = str(row['Mutual Fund'])
                    current_rank = row['Rank']
                    
                    # Try case-sensitive match first
                    if fund_name in reference_data[matching_sheet]:
                        old_rank = reference_data[matching_sheet][fund_name]
                        # Calculate rank change (as numeric)
                        rank_change = old_rank - current_rank
                        filtered_df.at[idx, 'Rank Change'] = rank_change
                    else:
                        # Try case-insensitive match as fallback
                        fund_name_lower = fund_name.lower()
                        for ref_name, ref_rank in reference_data[matching_sheet].items():
                            if ref_name.lower() == fund_name_lower:
                                # Calculate rank change (as numeric)
                                rank_change = ref_rank - current_rank
                                filtered_df.at[idx, 'Rank Change'] = rank_change
                                break

        # Append the 'total' rows at the bottom
        total_rows_with_change = total_rows[[fund_name_col, column]].copy()
        total_rows_with_change = total_rows_with_change.rename(columns={fund_name_col: 'Mutual Fund'})
        # Set Rank Change to empty for total rows
        total_rows_with_change['Rank Change'] = None
        filtered_df = pd.concat([filtered_df, total_rows_with_change])

        # Create a new sheet
        ws = wb.create_sheet(sheet_name)

        # Add a blank first column
        ws.insert_cols(1)

        # Write the DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(filtered_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 2):  # Start from column 2
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Format the first row (headers)
        header_fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6', fill_type='solid')
        header_font = Font(bold=True)
        for cell in ws[1]:
            if cell.column != 1:  # Skip the first column
                cell.fill = header_fill
                cell.font = header_font

        # Add borders to the table (excluding the first column)
        border = Side(border_style='thin', color='000000')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            for cell in row:
                cell.border = Border(left=border, right=border, top=border, bottom=border)

        # Format the last 'total' row in bold (excluding the first column)
        for cell in ws[ws.max_row]:
            if cell.column != 1:  # Skip the first column
                cell.font = Font(bold=True)

        # Highlight rows containing 'bajaj finserv' (case-insensitive) with light blue
        highlight_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
        highlight_font = Font(bold=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            if any(cell.value and 'bajaj finserv' in str(cell.value).lower() for cell in row):
                for cell in row:
                    cell.fill = highlight_fill
                    cell.font = highlight_font

        # Color code the Rank Change column
        rank_change_col = None
        for cell in ws[1]:  # Find the Rank Change column
            if cell.value == 'Rank Change':
                rank_change_col = cell.column
                break
        
        if rank_change_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=rank_change_col)
                # Skip if cell value is None (for total rows)
                if cell.value is not None:
                    # Format the cell based on its value
                    if cell.value > 0:
                        cell.font = Font(color='008800')  # Green for positive change
                        # Add up arrow for positive change
                        cell.value = f"{cell.value} ↑"
                    elif cell.value < 0:
                        cell.font = Font(color='CC0000')  # Red for negative change
                        # Add down arrow for negative change
                        cell.value = f"{cell.value} ↓"
                    elif cell.value == 0:
                        # Add neutral box for no change
                        cell.value = f"{cell.value} ■"

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

    # Remove the default sheet created by openpyxl
    wb.remove(wb['Sheet'])

    # Save the workbook
    wb.save(output_file)
    return True


def main():
    st.set_page_config(
        page_title="League Table Automation",
        page_icon="📊",
        layout="centered",
        initial_sidebar_state="expanded",
    )

    st.title("League Table Automation")
    st.markdown("""
    This application processes Excel files containing mutual fund data and generates a new file with:
    - Rank calculations for each fund
    - Rank change indicators with arrows (↑ for improvement, ↓ for decline, ■ for no change)
    - Formatted output with color-coding and highlighting
    """)

    st.header("Upload Files")
    
    input_file = st.file_uploader("Upload Input Excel File", type=["xlsx"], help="The input file should contain a 'MutualFund Name' column")
    reference_file = st.file_uploader("Upload Previous League Table File", type=["xlsx"], help="The previous league table is used to calculate rank changes")
    
    if st.button("Process Files", disabled=not input_file):
        if input_file:
            with st.spinner("Processing Excel files..."):
                # Create temporary files
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_input:
                    tmp_input.write(input_file.getvalue())
                    tmp_input_path = tmp_input.name
                
                tmp_ref_path = None
                if reference_file:
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_ref:
                        tmp_ref.write(reference_file.getvalue())
                        tmp_ref_path = tmp_ref.name
                
                # Create output file
                output_path = os.path.join(tempfile.gettempdir(), 'Output_File.xlsx')
                
                # Process the files
                success = process_excel(tmp_input_path, output_path, tmp_ref_path)
                
                # Clean up temporary input and reference files
                try:
                    os.unlink(tmp_input_path)
                except PermissionError:
                    pass
                
                if tmp_ref_path:
                    try:
                        os.unlink(tmp_ref_path)
                    except PermissionError:
                        pass
                
                if success:
                    # Read the output file for download
                    with open(output_path, "rb") as file:
                        output_data = file.read()
                    
                    # Clean up the output file
                    try:
                        os.unlink(output_path)
                    except PermissionError:
                        pass
                    
                    # Create download button
                    st.success("Processing complete! Click below to download the output file.")
                    st.download_button(
                        label="Download Output File",
                        data=output_data,
                        file_name="Output_File.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.info("""
                    **Output File Features:**
                    - Mutual funds ranked by value in descending order
                    - Rank change indicators with arrows (↑ for improvement, ↓ for decline, ■ for no change)
                    - Color-coded rank changes (green for improvements, red for declines)
                    - Highlighted rows containing "bajaj finserv"
                    - Bold formatting for total rows
                    """)
                else:
                    st.error("An error occurred during processing. Please check your input files.")
    
    st.sidebar.header("About")
    st.sidebar.info("""
    **LEAGUE TABLE AUTOMATION**
    
    This app processes mutual fund data from Excel files and generates a formatted output with rank calculations and change indicators.
    
    **Required Input Format:**
    - Input file must have a 'MutualFund Name' column
    - Previous month's league table is used to calculate rank changes
    - Make sure the D1 cell of the previous month's league table matches the column names in the input file
    
    **Output Features:**
    - Rank calculations
    - Rank change indicators with arrows
    - Formatted tables with color-coding
    """)


if __name__ == '__main__':
    main()